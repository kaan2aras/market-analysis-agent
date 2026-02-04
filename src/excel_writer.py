"""Excel writer with professional formatting and hyperlinks."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.config import FIELD_NAMES, FIELD_MAPPING, HEADER_COLOR, HEADER_FONT_COLOR, EVEN_ROW_COLOR, ODD_ROW_COLOR


class ExcelWriter:
    """Writes AI-powered app data to Excel with professional formatting."""
    
    def __init__(self, filename):
        """
        Initialize the Excel writer.
        
        Args:
            filename (str): Output Excel filename
        """
        self.filename = filename
        self.workbook = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
    
    def _create_header(self, sheet):
        """
        Create formatted header row.
        
        Args:
            sheet: Worksheet object
        """
        header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, field_name in enumerate(FIELD_NAMES, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = field_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
    
    def _format_cell(self, cell, value, is_link=False, row_num=2):
        """
        Format a cell with proper styling.
        
        Args:
            cell: Cell object
            value: Cell value
            is_link (bool): Whether the cell contains a hyperlink
            row_num (int): Row number for alternating colors
        """
        # Alternating row colors
        if row_num % 2 == 0:
            fill = PatternFill(start_color=EVEN_ROW_COLOR, end_color=EVEN_ROW_COLOR, fill_type="solid")
        else:
            fill = PatternFill(start_color=ODD_ROW_COLOR, end_color=ODD_ROW_COLOR, fill_type="solid")
        
        cell.fill = fill
        
        # Border
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = border
        
        # Alignment
        if is_link:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.font = Font(color="0563C1", underline="single")  # Blue, underlined like a link
        else:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        cell.value = value
    
    def _set_column_widths(self, sheet):
        """
        Set optimal column widths.
        
        Args:
            sheet: Worksheet object
        """
        column_widths = {
            "A": 20,  # App Name
            "B": 25,  # Company/Developer
            "C": 18,  # Category
            "D": 40,  # AI Features
            "E": 15,  # App Store Link
            "F": 15,  # Play Store Link
            "G": 15,  # Website
            "H": 12,  # Rating (iOS)
            "I": 12,  # Rating (Android)
            "J": 15,  # Downloads
            "K": 12,  # Pricing
            "L": 18,  # Monthly Active Users
            "M": 12,  # Trending Score
            "N": 20,  # Region Availability
            "O": 15   # Last Updated
        }
        
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
    
    def _write_apps_to_sheet(self, sheet, apps, sheet_name):
        """
        Write app data to a worksheet.
        
        Args:
            sheet: Worksheet object
            apps (list): List of app dictionaries
            sheet_name (str): Name of the worksheet
        """
        sheet.title = sheet_name
        
        # Create header
        self._create_header(sheet)
        
        # Write data
        for row_num, app in enumerate(apps, start=2):
            for col_num, (db_key, field_name) in enumerate(FIELD_MAPPING.items(), start=1):
                cell = sheet.cell(row=row_num, column=col_num)
                value = app.get(db_key, "")
                
                # Check if this is a link column
                is_link = db_key in ["app_store_link", "play_store_link", "website"]
                
                if is_link and value:
                    # Create hyperlink
                    cell.hyperlink = value
                    cell.value = "View Link"
                    self._format_cell(cell, "View Link", is_link=True, row_num=row_num)
                else:
                    self._format_cell(cell, value, is_link=False, row_num=row_num)
        
        # Set column widths
        self._set_column_widths(sheet)
        
        # Freeze top row
        sheet.freeze_panes = "A2"
    
    def write_all_apps(self, apps):
        """
        Create 'All Apps' sheet.
        
        Args:
            apps (list): List of all apps
        """
        sheet = self.workbook.create_sheet("All Apps")
        self._write_apps_to_sheet(sheet, apps, "All Apps")
    
    def write_by_category(self, categories_dict):
        """
        Create 'By Category' sheet with apps grouped by category.
        
        Args:
            categories_dict (dict): Dictionary with categories as keys
        """
        sheet = self.workbook.create_sheet("By Category")
        self._create_header(sheet)
        
        row_num = 2
        for category, apps in sorted(categories_dict.items()):
            # Add category header row
            category_cell = sheet.cell(row=row_num, column=1)
            category_cell.value = f"▼ {category}"
            category_cell.font = Font(bold=True, size=12, color="000000")
            category_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            
            # Merge cells for category header
            sheet.merge_cells(f"A{row_num}:O{row_num}")
            row_num += 1
            
            # Add apps in this category
            for app in apps:
                for col_num, (db_key, field_name) in enumerate(FIELD_MAPPING.items(), start=1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = app.get(db_key, "")
                    
                    is_link = db_key in ["app_store_link", "play_store_link", "website"]
                    
                    if is_link and value:
                        cell.hyperlink = value
                        cell.value = "View Link"
                        self._format_cell(cell, "View Link", is_link=True, row_num=row_num)
                    else:
                        self._format_cell(cell, value, is_link=False, row_num=row_num)
                
                row_num += 1
        
        self._set_column_widths(sheet)
        sheet.freeze_panes = "A2"
    
    def write_top_trending(self, apps):
        """
        Create 'Top Trending' sheet.
        
        Args:
            apps (list): List of trending apps sorted by score
        """
        sheet = self.workbook.create_sheet("Top Trending")
        self._write_apps_to_sheet(sheet, apps, "Top Trending")
    
    def write_by_region(self, regions_dict):
        """
        Create 'By Region' sheet with apps grouped by region.
        
        Args:
            regions_dict (dict): Dictionary with regions as keys
        """
        sheet = self.workbook.create_sheet("By Region")
        self._create_header(sheet)
        
        row_num = 2
        for region, apps in sorted(regions_dict.items()):
            # Add region header row
            region_cell = sheet.cell(row=row_num, column=1)
            region_cell.value = f"▼ {region}"
            region_cell.font = Font(bold=True, size=12, color="000000")
            region_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            
            # Merge cells for region header
            sheet.merge_cells(f"A{row_num}:O{row_num}")
            row_num += 1
            
            # Add apps in this region
            for app in apps:
                for col_num, (db_key, field_name) in enumerate(FIELD_MAPPING.items(), start=1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = app.get(db_key, "")
                    
                    is_link = db_key in ["app_store_link", "play_store_link", "website"]
                    
                    if is_link and value:
                        cell.hyperlink = value
                        cell.value = "View Link"
                        self._format_cell(cell, "View Link", is_link=True, row_num=row_num)
                    else:
                        self._format_cell(cell, value, is_link=False, row_num=row_num)
                
                row_num += 1
        
        self._set_column_widths(sheet)
        sheet.freeze_panes = "A2"
    
    def save(self):
        """Save the workbook to file."""
        self.workbook.save(self.filename)
        print(f"Excel file saved: {self.filename}")
